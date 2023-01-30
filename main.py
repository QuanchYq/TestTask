import os
import openpyxl
#pip install openpyxl



def write_to_excel(root, name, ext, row, sheet):
    sheet.cell(row=row, column=1, value=row-1)
    sheet.cell(row=row, column=2, value=root)
    sheet.cell(row=row, column=3, value=name)
    sheet.cell(row=row, column=4, value=ext)

def get_file_info(root, filename):
    name, ext = os.path.splitext(filename)
    return root, name, ext

def main():
    # Initialize the workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value='Line number')
    sheet.cell(row=1, column=2, value='Folder containing the file')
    sheet.cell(row=1, column=3, value='File name')
    sheet.cell(row=1, column=4, value='File extension')

    # Set the starting row
    row = 2

    # Start looping through the folder and its subfolders
    for root, dirs, files in os.walk('C:/Users/aikyn/Desktop/testtask'):
        for filename in files:
            root, name, ext = get_file_info(root, filename)
            write_to_excel(root, name, ext, row, sheet)
            row += 1

    # Save the workbook
    workbook.save('result.xlsx')

if __name__ == '__main__':
    main()
    print('Successfully')

